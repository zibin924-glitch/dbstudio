"""Query result export service - CSV, Excel, and JSON formats."""

import csv
import io
import json
import logging
from typing import Any

logger = logging.getLogger(__name__)


class ExportService:
    """Exports query results to various file formats (CSV, Excel, JSON)."""

    @staticmethod
    def export_csv(columns: list[str], rows: list[list[Any]]) -> bytes:
        """Generate a CSV file from query results.

        Args:
            columns: List of column name strings.
            rows: List of row data (each row is a list of values).

        Returns:
            CSV content as bytes (UTF-8 with BOM for Excel compatibility).
        """
        output = io.StringIO()
        writer = csv.writer(output, quoting=csv.QUOTE_MINIMAL)

        # Write header
        writer.writerow(columns)

        # Write data rows
        for row in rows:
            # Convert None to empty string and stringify other values
            cleaned = []
            for val in row:
                if val is None:
                    cleaned.append("")
                else:
                    cleaned.append(str(val))
            writer.writerow(cleaned)

        csv_text = output.getvalue()
        output.close()

        # UTF-8 BOM for Excel compatibility
        bom = b"\xef\xbb\xbf"
        return bom + csv_text.encode("utf-8")

    @staticmethod
    def export_excel(columns: list[str], rows: list[list[Any]]) -> bytes:
        """Generate an Excel (.xlsx) file from query results using openpyxl.

        Args:
            columns: List of column name strings.
            rows: List of row data (each row is a list of values).

        Returns:
            Excel file content as bytes.
        """
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Font, PatternFill

        wb = Workbook()
        ws = wb.active
        ws.title = "Query Results"

        # Style for header row
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_alignment = Alignment(horizontal="center")

        # Write header
        for col_idx, col_name in enumerate(columns, start=1):
            cell = ws.cell(row=1, column=col_idx, value=col_name)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment

        # Write data rows
        for row_idx, row in enumerate(rows, start=2):
            for col_idx, value in enumerate(row, start=1):
                cell = ws.cell(row=row_idx, column=col_idx)
                if value is None:
                    cell.value = None
                elif isinstance(value, (int, float)):
                    cell.value = value
                else:
                    cell.value = str(value)

        # Auto-adjust column widths (approximate)
        for col_idx, col_name in enumerate(columns, start=1):
            max_len = len(str(col_name))
            for row in rows:
                if col_idx - 1 < len(row) and row[col_idx - 1] is not None:
                    max_len = max(max_len, len(str(row[col_idx - 1])))
            # Cap at 50 characters
            ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = min(
                max_len + 4, 50
            )

        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        return output.read()

    @staticmethod
    def export_json(columns: list[str], rows: list[list[Any]]) -> list[dict]:
        """Convert query results to a list of dictionaries.

        Each dictionary maps column names to their values for one row.

        Args:
            columns: List of column name strings.
            rows: List of row data.

        Returns:
            List of row dictionaries.
        """
        result = []
        for row in rows:
            entry = {}
            for col_idx, col_name in enumerate(columns):
                value = row[col_idx] if col_idx < len(row) else None
                # Convert non-serializable types to strings
                if value is not None and not isinstance(value, (str, int, float, bool)):
                    value = str(value)
                entry[col_name] = value
            result.append(entry)
        return result
